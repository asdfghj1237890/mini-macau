"""
Build public/data/schools.json: every Macau school / tertiary campus with its
building footprints, so the map can draw them as coloured 3D blocks by level.

Why a separate building layer instead of tinting the basemap: OpenFreeMap's
`building` tiles merge all buildings of the same height into one multipolygon
feature (a z14 tile holds ~8,000 footprints in ~120 features), so per-building
feature-state is impossible for anything but isolated skyscrapers. We therefore
take the footprints straight from OpenStreetMap and render our own extrusions.

Sources
  * DSEDJ 非高等教育學校清單 on data.gov.mo (77 schools with approved levels:
    幼稚園 / 小學 / 中學). This decides the colour class:
      all three  -> all_through (一條龍)
      has 中     -> secondary
      has 小     -> primary
      has 幼     -> kindergarten
      回歸教育 night schools (no flags) -> secondary
  * OpenStreetMap via Overpass: amenity=school/kindergarten/college/university
    features (campus polygons or nodes) + the building ways inside them.
    Tertiary campuses (amenity=university, or college whose name says
    大學/學院/高等) become `university`. OSM education features that match no
    DSEDJ school and are not tertiary are dropped (training centres, creches).

Matching DSEDJ <-> OSM is by name: exact / substring on the Chinese name (OSM
names are campus-level, e.g. 勞校中學附屬小學 ⊂ 勞校中學), then a strict
Portuguese-name ratio, plus a small alias table for the known odd ones.

Run manually when the school list or OSM changes (not scheduled):
    cd data && uv run python scripts/fetch_schools.py
Needs network (data.gov.mo + overpass-api.de); ~25 Overpass calls. Budget
5–15 minutes: overpass-api.de answers 429 to back-to-back calls and the
exponential backoff (5 s … 160 s) dominates the run time.
"""

import difflib
import hashlib
import io
import json
import math
import re
import sys
import tempfile
import time
import unicodedata
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import mapbox_vector_tile as mvt
import requests
from openpyxl import load_workbook
from shapely.geometry import LineString, MultiPolygon, Point, Polygon
from shapely.ops import polygonize, unary_union

ROOT = Path(__file__).parent.parent.parent
OUTPUT_PATH = ROOT / "public" / "data" / "schools.json"
BUS_STOPS_PATH = ROOT / "public" / "data" / "bus-stops.json"

DSEDJ_DATASET_ID = "f0578833-7dd6-4ed5-b825-75e9c4f56012"
DSEDJ_PAGE = f"https://data.gov.mo/Detail?id={DSEDJ_DATASET_ID}"
DSEDJ_DOWNLOAD = f"https://api.data.gov.mo/document/download/{DSEDJ_DATASET_ID}"
# Tried in order; a connection failure / 5xx moves to the next mirror for that
# call (429 = rate limit, retried on the same host). All run the same API;
# the mirrors can lag the main instance by minutes, which does not matter here.
OVERPASS_ENDPOINTS = (
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
)
# south, west, north, east — generous; the bus-stop proximity test below trims
# the Zhuhai / Hengqin side that this box unavoidably includes.
BBOX = "22.06,113.52,22.23,113.62"
MACAU_MAX_STOP_DISTANCE_M = 500
HEADERS = {"User-Agent": "mini-macau data pipeline (https://github.com/asdfghj1237890/mini-macau)"}

LEVELS = ("kindergarten", "primary", "secondary", "university", "all_through")
DEFAULT_RENDER_HEIGHT_M = 5.0  # OpenMapTiles default for untagged buildings
LEVEL_HEIGHT_M = 3.66  # OpenMapTiles building:levels -> metres
BUILDING_BUFFER_M = 0.5  # grow footprints so our walls do not z-fight the basemap's
# Heights are stored as the basemap draws them (no margin). The frontend adds
# its own vertical margin when it builds the layer (see SCHOOL_HEIGHT_MARGIN_M
# in src/schools.ts): a data-side 0.5 m was not enough — large low roofs
# z-fought the basemap's roof at oblique angles, and the z14→15.5 height ramp
# shrank the margin further.
EXTRA_HEIGHT_M = 0.0
NODE_SEARCH_RADIUS_M = 25
POLY_CHUNK = 20  # school polygons per Overpass request
TERTIARY_NAME_RE = re.compile(r"大學|學院|高等|Universidade|Instituto|University")
# OSM campus polygons are sometimes drawn around a whole block; buildings that
# are clearly not the school's (flats, hotels, offices, car parks) are left to
# the basemap even when their centroid falls inside the campus outline.
FOREIGN_BUILDING_KINDS = {
    "apartments", "residential", "house", "detached", "terrace",
    "hotel", "commercial", "retail", "office", "industrial", "warehouse",
    "garage", "garages", "parking", "supermarket", "temple", "church",
}

# OSM name (normalised, see norm()) -> DSEDJ school name (normalised).
ALIASES = {
    "st.anthony's": "聖安東尼幼稚園",
    "stanthony's": "聖安東尼幼稚園",
    "培道學校氹仔分校": "培道中學",
    "聖若瑟教區中學校第五校": "聖若瑟教區中學第五校",
    "勞工子弟學校": "勞校中學",  # full name of the school known as 勞校
    "同善堂小學": "同善堂中學",  # primary section of the all-through 同善堂中學
}

MIN_SCHOOLS = 40  # degenerate-run guard

# Fallback footprint source: the very tiles the map draws. When OSM has no
# `building` way inside a campus (buildings mapped as relations, building:part,
# or not at all), we cut the basemap's own merged building multipolygons by the
# campus outline so our colour covers exactly what the basemap renders there.
OFM_TILEJSON = "https://tiles.openfreemap.org/planet"
TILE_ZOOM = 14  # OpenFreeMap maxzoom; higher zooms are overzoomed from this


# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
def norm(s: str | None, strip_macau: bool = True) -> str:
    s = unicodedata.normalize("NFKC", s or "")
    s = re.sub(r"[\s\-–—·•,，。、()（）\[\]「」『』/]", "", s)
    return s.replace("澳門", "").replace("澳门", "") if strip_macau else s


# Substring matching on a name that is only generic words once 澳門 is
# stripped ("國際學校", "中葡學校") would glue unrelated schools together
# (澳門國際學校 ≠ 傳承國際學校), so the stripped form needs this many chars.
MIN_STRIPPED_SUBSTR_LEN = 5


def zh_part(s: str) -> str:
    return "".join(ch for ch in (s or "") if "一" <= ch <= "鿿")


def latin_part(s: str) -> str:
    return re.sub(r"[一-鿿（）]", "", s or "").strip()


def http_get(url: str, attempts: int = 6, **kw) -> requests.Response:
    last = None
    for i in range(attempts):
        try:
            r = requests.get(url, headers=HEADERS, timeout=60, **kw)
            if r.status_code == 200:
                return r
            last = f"HTTP {r.status_code}"
        except requests.RequestException as e:
            last = f"{type(e).__name__}: {e}"
        time.sleep(2 * (2**i))
    raise RuntimeError(f"GET {url} failed: {last}")


# Overpass answers are cached in the OS temp dir for a day, keyed by the query
# text: a run that dies half-way (this machine's network drops; overpass-api.de
# 429s back-to-back calls) can be restarted without repeating finished calls.
OVERPASS_CACHE_DIR = Path(tempfile.gettempdir()) / "mini-macau-overpass-cache"
OVERPASS_CACHE_TTL_S = 24 * 3600
# Index of the mirror that last answered; sticky across calls so one dead host
# costs its connect timeout once per run, not once per query.
_overpass_endpoint = 0


def overpass(query: str, attempts: int = 8) -> list[dict]:
    """POST an Overpass QL query; backs off on 429/5xx and transient errors."""
    global _overpass_endpoint
    OVERPASS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file = OVERPASS_CACHE_DIR / (hashlib.sha1(query.encode("utf-8")).hexdigest() + ".json")
    if cache_file.exists() and time.time() - cache_file.stat().st_mtime < OVERPASS_CACHE_TTL_S:
        return json.loads(cache_file.read_text(encoding="utf-8"))
    last = None
    same_host_failures = 0
    for i in range(attempts):
        url = OVERPASS_ENDPOINTS[_overpass_endpoint % len(OVERPASS_ENDPOINTS)]
        try:
            r = requests.post(url, data={"data": query}, headers=HEADERS, timeout=(20, 180))
            if r.status_code == 200:
                elements = r.json().get("elements", [])
                cache_file.write_text(json.dumps(elements), encoding="utf-8")
                time.sleep(2)  # be polite between calls
                return elements
            last = f"HTTP {r.status_code} from {url}"
            if r.status_code == 429:
                same_host_failures += 1  # rate limit: back off on this host
            else:
                _overpass_endpoint += 1  # server-side trouble: next mirror
                same_host_failures = 0
        except (requests.RequestException, ValueError) as e:
            last = f"{type(e).__name__} from {url}: {str(e)[:120]}"
            _overpass_endpoint += 1  # unreachable host: next mirror
            same_host_failures = 0
        # Short pause right after switching mirrors; exponential (capped)
        # only while hammering the same host.
        delay = min(5 * (2**same_host_failures), 60)
        print(f"  overpass attempt {i + 1} failed ({last}); retrying in {delay}s", file=sys.stderr)
        time.sleep(delay)
    raise RuntimeError(f"Overpass failed: {last}")


def metres_xy(lng: float, lat: float, lat0: float) -> tuple[float, float]:
    return (lng * 111320.0 * math.cos(math.radians(lat0)), lat * 110540.0)


def xy_lnglat(x: float, y: float, lat0: float) -> tuple[float, float]:
    return (x / (111320.0 * math.cos(math.radians(lat0))), y / 110540.0)


def ring_from_geometry(geom: list[dict]) -> list[tuple[float, float]]:
    return [(p["lon"], p["lat"]) for p in geom]


def polygon_of_element(el: dict) -> Polygon | MultiPolygon | None:
    """Campus polygon for an OSM way (closed) or relation (outer members)."""
    if el["type"] == "way":
        ring = ring_from_geometry(el.get("geometry", []))
        if len(ring) >= 4 and ring[0] == ring[-1]:
            poly = Polygon(ring)
            return poly if poly.is_valid else poly.buffer(0)
        return None
    if el["type"] == "relation":
        lines = [
            LineString(ring_from_geometry(m["geometry"]))
            for m in el.get("members", [])
            if m.get("role") == "outer" and m.get("geometry") and len(m["geometry"]) >= 2
        ]
        if not lines:
            return None
        polys = list(polygonize(unary_union(lines)))
        if not polys:
            return None
        return unary_union(polys)
    return None


def parse_height(tags: dict) -> tuple[float, float]:
    """(render_height, render_min_height) following the OpenMapTiles rule."""

    def num(v: str | None) -> float | None:
        if not v:
            return None
        m = re.match(r"^\s*(-?\d+(?:\.\d+)?)", v.replace(",", "."))
        return float(m.group(1)) if m else None

    h = num(tags.get("height"))
    if h is None:
        lv = num(tags.get("building:levels"))
        h = lv * LEVEL_HEIGHT_M if lv else DEFAULT_RENDER_HEIGHT_M
    mh = num(tags.get("min_height"))
    if mh is None:
        ml = num(tags.get("building:min_level"))
        mh = ml * LEVEL_HEIGHT_M if ml else 0.0
    # The tiles carry integer heights: planetiler rounds `height` UP (11.3 → 12,
    # 23.1 → 24, verified against MUST's buildings), so our top must clear the
    # rounded-up value or the basemap's grey roof pokes through our colour.
    # The base is rounded down for the same reason (we must start no higher).
    return (float(math.ceil(h - 1e-9)), float(math.floor(mh + 1e-9)))


def buffered_footprint(poly: Polygon, lat0: float) -> list[list[list[float]]]:
    """Outer ring(s) grown by BUILDING_BUFFER_M, as GeoJSON polygon coordinates."""
    xy = Polygon([metres_xy(x, y, lat0) for x, y in poly.exterior.coords])
    grown = xy.buffer(BUILDING_BUFFER_M, join_style="mitre").simplify(0.3)
    if grown.geom_type == "MultiPolygon":
        grown = max(grown.geoms, key=lambda g: g.area)
    ring = [list(map(lambda v: round(v, 6), xy_lnglat(x, y, lat0))) for x, y in grown.exterior.coords]
    return [ring]


# ----------------------------------------------------------------------------
# basemap tiles (fallback footprints)
# ----------------------------------------------------------------------------
def tile_xy(lng: float, lat: float, z: int) -> tuple[int, int]:
    n = 2**z
    x = int((lng + 180.0) / 360.0 * n)
    lat_r = math.radians(lat)
    y = int((1.0 - math.log(math.tan(lat_r) + 1.0 / math.cos(lat_r)) / math.pi) / 2.0 * n)
    return x, y


def tile_px_to_lnglat(px: float, py: float, tx: int, ty: int, z: int, extent: int) -> tuple[float, float]:
    n = 2**z
    lng = (tx + px / extent) / n * 360.0 - 180.0
    lat = math.degrees(math.atan(math.sinh(math.pi * (1.0 - 2.0 * (ty + py / extent) / n))))
    return (lng, lat)


def fetch_tile_building_parts(tiles: set[tuple[int, int]]) -> list[dict]:
    """Every building polygon part in the given z14 tiles, in lng/lat, with the
    basemap's render heights. Merged multipolygons are split into their parts."""
    tilejson = http_get(OFM_TILEJSON).json()
    template = tilejson["tiles"][0]
    parts: list[dict] = []
    seen_shapes: set[tuple] = set()  # the tiles repeat some footprints verbatim
    for tx, ty in sorted(tiles):
        url = template.replace("{z}", str(TILE_ZOOM)).replace("{x}", str(tx)).replace("{y}", str(ty))
        data = http_get(url).content
        if not data:
            continue
        layer = mvt.decode(data, default_options={"y_coord_down": True}).get("building")
        if not layer:
            continue
        extent = layer.get("extent", 4096)
        for feat in layer["features"]:
            geom = feat["geometry"]
            polys = geom["coordinates"] if geom["type"] == "MultiPolygon" else [geom["coordinates"]]
            props = feat.get("properties", {})
            for i, rings in enumerate(polys):
                if not rings or len(rings[0]) < 4:
                    continue
                ring = [tile_px_to_lnglat(px, py, tx, ty, TILE_ZOOM, extent) for px, py in rings[0]]
                poly = Polygon(ring)
                if not poly.is_valid:
                    poly = poly.buffer(0)
                if poly.is_empty or poly.geom_type != "Polygon":
                    continue
                shape_key = (round(poly.centroid.x, 6), round(poly.centroid.y, 6), round(poly.area * 1e10), props.get("render_height"))
                if shape_key in seen_shapes:
                    continue
                seen_shapes.add(shape_key)
                parts.append(
                    {
                        "_poly": poly,
                        "id": f"t{TILE_ZOOM}/{tx}/{ty}/{feat.get('id', 0)}#{i}",
                        "height": float(props.get("render_height") or DEFAULT_RENDER_HEIGHT_M),
                        "minHeight": float(props.get("render_min_height") or 0.0),
                    }
                )
        print(f"  tile {TILE_ZOOM}/{tx}/{ty}: {len(parts)} building parts so far")
    return parts


# ----------------------------------------------------------------------------
# DSEDJ school list
# ----------------------------------------------------------------------------
def fetch_dsedj_schools() -> list[dict]:
    print(f"Fetching DSEDJ school list {DSEDJ_DOWNLOAD}")
    blob = None
    for i in range(8):
        r = http_get(DSEDJ_DOWNLOAD)
        if r.content[:2] == b"PK":
            blob = r.content
            break
        print(f"  non-ZIP body ({r.content[:60]!r}); retrying", file=sys.stderr)
        time.sleep(3)
    if blob is None:
        raise RuntimeError("DSEDJ download never returned a ZIP")
    z = zipfile.ZipFile(io.BytesIO(blob))
    name = next(n for n in z.namelist() if n.lower().endswith(".xlsx"))
    ws = load_workbook(io.BytesIO(z.read(name)), data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        rec = dict(zip(hdr, r))
        if not rec.get("s_code"):
            continue
        flags = {
            "kindergarten": bool(rec.get("approv_i")),
            "primary": bool(rec.get("approv_p")),
            "secondary": bool(rec.get("approv_sec")),
        }
        edutype = str(rec.get("edutype") or "")
        if flags["kindergarten"] and flags["primary"] and flags["secondary"]:
            level = "all_through"
        elif flags["secondary"]:
            level = "secondary"
        elif flags["primary"]:
            level = "primary"
        elif flags["kindergarten"]:
            level = "kindergarten"
        else:
            level = "secondary"  # 回歸教育 evening schools carry no flags
        out.append(
            {
                "code": str(rec["s_code"]).strip(),
                "nameZh": str(rec.get("s_code_name_c") or "").strip(),
                "namePt": re.sub(r"\s+", " ", str(rec.get("s_code_name_p") or "")).strip(),
                "system": "public" if "公立" in str(rec.get("School_sys") or "") else "private",
                "level": level,
                "levels": flags,
                "eveningOnly": "回歸" in edutype and "正規" not in edutype,
            }
        )
    print(f"  {len(out)} schools ({sum(1 for s in out if s['level'] == 'all_through')} all-through)")
    return out


# ----------------------------------------------------------------------------
# OSM education features
# ----------------------------------------------------------------------------
def fetch_osm_education() -> list[dict]:
    print("Fetching OSM education features (Overpass)")
    # `out geom` (not `out tags geom`): the tags-only mode drops relation
    # members, and the university campuses are multipolygon relations.
    # Many Macau schools are mapped only as a named building (building=school
    # / building=kindergarten / building=yes + "…學校"), with no amenity tag;
    # pull those in too — the DSEDJ name matching decides what is a school,
    # so a block of flats called 海暉閣 is still dropped.
    els = overpass(
        f'[out:json][timeout:120][bbox:{BBOX}];'
        '(nwr["amenity"="school"];nwr["amenity"="kindergarten"];'
        'nwr["amenity"="college"];nwr["amenity"="university"];'
        'nwr["building"="school"];nwr["building"="kindergarten"];'
        'nwr["building"]["name"~"學校|中學|小學|幼稚園|書院"];);out geom;'
    )
    print(f"  {len(els)} features in bbox")
    return els


def load_bus_stop_points() -> list[tuple[float, float]]:
    pts: list[tuple[float, float]] = []

    def walk(o):
        if isinstance(o, dict):
            c = o.get("coordinates")
            if isinstance(c, list) and len(c) == 2 and all(isinstance(v, (int, float)) for v in c):
                pts.append((c[0], c[1]))
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)

    walk(json.loads(BUS_STOPS_PATH.read_text(encoding="utf-8")))
    return pts


def representative_point(el: dict) -> tuple[float, float]:
    if el["type"] == "node":
        return (el["lon"], el["lat"])
    poly = polygon_of_element(el)
    if poly is not None and not poly.is_empty:
        p = poly.representative_point()
        return (p.x, p.y)
    c = el.get("center") or {}
    return (c.get("lon", 0.0), c.get("lat", 0.0))


# ----------------------------------------------------------------------------
# matching
# ----------------------------------------------------------------------------
def match_schools(dsedj: list[dict], osm: list[dict]) -> tuple[dict[str, list[dict]], list[dict]]:
    """Returns ({dsedj code: [osm feature, ...]}, [unmatched osm features])."""
    by_code: dict[str, list[dict]] = {s["code"]: [] for s in dsedj}
    unmatched: list[dict] = []
    for el in osm:
        tags = el["tags"]
        name = tags.get("name", "")
        zh = tags.get("name:zh") or tags.get("name:zh-Hant") or zh_part(name)
        pt = tags.get("name:pt") or latin_part(name)
        nzh = norm(zh)
        nzh_full = norm(zh, strip_macau=False)
        npt = norm(pt).lower()
        alias = ALIASES.get(norm(name).lower()) or ALIASES.get(nzh)
        cands: list[tuple[float, str, dict]] = []
        for s in dsedj:
            snzh = norm(s["nameZh"])
            snzh_full = norm(s["nameZh"], strip_macau=False)
            snpt = norm(s["namePt"]).lower()
            score, how = 0.0, ""
            if alias and norm(alias) == snzh:
                score, how = 1.0, "alias"
            elif nzh and snzh:
                shorter = min(len(nzh), len(snzh))
                contains_full = snzh_full in nzh_full or nzh_full in snzh_full
                contains_stripped = snzh in nzh or nzh in snzh
                if nzh == snzh:
                    score, how = 1.0, "exact"
                elif shorter >= 3 and (contains_full or (contains_stripped and shorter >= MIN_STRIPPED_SUBSTR_LEN)):
                    # campus suffixes (附屬小學, 路環校部, (第五校)) or DSEDJ prefixes
                    # (澳門…). Longer DSEDJ names win ties, so 聖若瑟教區中學第五校
                    # beats 聖若瑟教區中學 for the 第五校 campus.
                    score, how = 0.9 + len(snzh) / 1000, "substr"
                else:
                    ratio = difflib.SequenceMatcher(None, nzh, snzh).ratio()
                    if ratio >= 0.85:
                        score, how = 0.8 * ratio, f"fuzzy{ratio:.2f}"
            if score < 0.8 and not nzh and npt and snpt:
                # English/Portuguese-only OSM names: strict Latin match only
                ratio = difflib.SequenceMatcher(None, npt, snpt).ratio()
                if ratio >= 0.95:
                    score, how = 0.85, f"pt{ratio:.2f}"
            if score > 0:
                cands.append((score, how, s))
        if not cands:
            unmatched.append(el)
            continue
        cands.sort(key=lambda c: -c[0])
        best = cands[0]
        # An OSM feature names several schools when they share a campus
        # (e.g. "嘉諾撒聖心中學 / 嘉諾撒聖心英文中學"): keep every DSEDJ school
        # whose name is spelled out inside the OSM name and is not itself
        # contained in a better winner. The reverse containment (a DSEDJ name
        # that merely starts with the OSM name, e.g. 濠江中學附屬橫琴學校 vs
        # the OSM 濠江中學 campus) is a different school and must not ride along.
        winners = [best]
        for score, how, s in cands[1:]:
            if how != "substr" or best[1] not in ("substr", "exact", "alias"):
                continue
            snzh = norm(s["nameZh"])
            if snzh not in nzh:
                continue
            if any(snzh in norm(w[2]["nameZh"]) for w in winners):
                continue
            winners.append((score, how, s))
        el["_match"] = [{"code": w[2]["code"], "how": w[1], "score": round(w[0], 3)} for w in winners]
        for w in winners:
            by_code[w[2]["code"]].append(el)
    return by_code, unmatched


# ----------------------------------------------------------------------------
# buildings
# ----------------------------------------------------------------------------
def fetch_buildings_for_polygons(polys: list[tuple[str, Polygon | MultiPolygon]]) -> list[dict]:
    """All building ways intersecting any of the campus polygons (chunked)."""
    found: dict[int, dict] = {}
    for i in range(0, len(polys), POLY_CHUNK):
        chunk = polys[i : i + POLY_CHUNK]
        parts = []
        for _, poly in chunk:
            geoms = list(poly.geoms) if poly.geom_type == "MultiPolygon" else [poly]
            for g in geoms:
                ring = " ".join(f"{y:.6f} {x:.6f}" for x, y in g.exterior.coords)
                parts.append(f'way["building"](poly:"{ring}");')
        els = overpass(f"[out:json][timeout:180];({''.join(parts)});out tags geom;")
        for el in els:
            found[el["id"]] = el
        print(f"  polygons {i + 1}-{i + len(chunk)}: {len(els)} building ways")
    return list(found.values())


def fetch_buildings_near_node(lng: float, lat: float) -> list[dict]:
    return overpass(f'[out:json][timeout:60];way["building"](around:{NODE_SEARCH_RADIUS_M},{lat:.7f},{lng:.7f});out tags geom;')


def fetch_buildings_in_relation(rel_id: int) -> list[dict]:
    return overpass(f'[out:json][timeout:180];way["building"](area:{3600000000 + rel_id});out tags geom;')


def building_record(el: dict, lat0: float) -> dict | None:
    ring = ring_from_geometry(el.get("geometry", []))
    if len(ring) < 4 or ring[0] != ring[-1]:
        return None
    poly = Polygon(ring)
    if not poly.is_valid:
        poly = poly.buffer(0)
    if poly.is_empty or poly.geom_type != "Polygon":
        return None
    h, mh = parse_height(el["tags"])
    return {
        "osmId": f"w{el['id']}",
        "name": el["tags"].get("name") or None,
        "kind": el["tags"].get("building") or "yes",
        "height": round(h + EXTRA_HEIGHT_M, 1),
        "minHeight": mh,
        "coordinates": buffered_footprint(poly, lat0),
        "_poly": poly,
    }


# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------
def run() -> int:
    dsedj = fetch_dsedj_schools()
    osm_all = fetch_osm_education()

    # Keep only features that are actually in Macau (bbox spills into Zhuhai).
    stops = load_bus_stop_points()
    lat0 = 22.16

    def near_macau(lng: float, lat: float) -> bool:
        x, y = metres_xy(lng, lat, lat0)
        return any(math.hypot(x - sx, y - sy) <= MACAU_MAX_STOP_DISTANCE_M for sx, sy in (metres_xy(a, b, lat0) for a, b in stops))

    osm = []
    for el in osm_all:
        lng, lat = representative_point(el)
        if near_macau(lng, lat):
            el["_point"] = (lng, lat)
            osm.append(el)
    print(f"  {len(osm)} inside Macau (bus-stop proximity filter)")

    by_code, unmatched_osm = match_schools(dsedj, osm)

    # Tertiary campuses come from OSM alone.
    tertiary = []
    dropped = []
    for el in unmatched_osm:
        am = el["tags"].get("amenity") or ""
        bld = el["tags"].get("building") or ""
        name = el["tags"].get("name", "")
        # amenity=university outright; amenity/building=college|university only
        # when the name says it is a tertiary institution (keeps 消防學校 out).
        if am == "university" or (
            (am == "college" or bld in ("university", "college")) and TERTIARY_NAME_RE.search(name)
        ):
            tertiary.append(el)
        else:
            dropped.append(el)
    print(f"  matched OSM features: {sum(len(v) for v in by_code.values())}; tertiary: {len(tertiary)}; dropped: {len(dropped)}")

    # --- collect campus geometries -------------------------------------------
    # entries: (owner key, osm element)
    owners: list[tuple[str, dict]] = []
    for code, els in by_code.items():
        for el in els:
            owners.append((f"dsedj:{code}", el))
    for el in tertiary:
        owners.append((f"osm:{el['type'][0]}{el['id']}", el))

    polygons: list[tuple[str, Polygon | MultiPolygon]] = []
    nodes: list[tuple[str, dict]] = []
    relations: list[tuple[str, dict]] = []
    self_buildings: list[tuple[str, dict]] = []
    for key, el in owners:
        if el["type"] == "node":
            nodes.append((key, el))
            continue
        poly = polygon_of_element(el)
        if poly is None or poly.is_empty:
            continue
        if el["type"] == "relation":
            relations.append((key, el))  # area query as well as the polygon query below
        if el["tags"].get("building"):
            self_buildings.append((key, el))  # the school outline IS the building
        polygons.append((key, poly))

    print(f"Fetching buildings: {len(polygons)} campus polygons, {len(nodes)} nodes, {len(relations)} relations")
    campus_buildings = fetch_buildings_for_polygons(polygons) if polygons else []

    # Assign each building way to exactly one owner: the polygon that contains
    # its centroid (first owner in DSEDJ order wins on shared campuses).
    claimed: dict[str, dict] = {}
    assigned: dict[str, list[dict]] = {}

    def claim(key: str, rec: dict) -> None:
        if rec["osmId"] in claimed:
            return
        claimed[rec["osmId"]] = rec
        assigned.setdefault(key, []).append(rec)

    for key, el in self_buildings:
        if el["type"] == "relation":
            # multipolygon building (courtyard school): extrude the outer ring(s)
            poly = polygon_of_element(el)
            geoms = list(poly.geoms) if poly.geom_type == "MultiPolygon" else [poly]
            h, mh = parse_height(el["tags"])
            for i, g in enumerate(geoms):
                claim(key, {
                    "osmId": f"r{el['id']}" + (f"#{i}" if i else ""),
                    "name": el["tags"].get("name") or None,
                    "kind": el["tags"].get("building") or "yes",
                    "height": round(h + EXTRA_HEIGHT_M, 1),
                    "minHeight": mh,
                    "coordinates": buffered_footprint(g, lat0),
                    "_poly": g,
                })
            continue
        rec = building_record(el, lat0)
        if rec:
            claim(key, rec)
    for el in campus_buildings:
        rec = building_record(el, lat0)
        if not rec or rec["kind"] in FOREIGN_BUILDING_KINDS:
            continue
        c = rec["_poly"].representative_point()
        for key, poly in polygons:
            if poly.contains(c):
                claim(key, rec)
                break
    for key, el in nodes:
        lng, lat = el["_point"]
        cands = [building_record(b, lat0) for b in fetch_buildings_near_node(lng, lat)]
        cands = [c for c in cands if c]
        if not cands:
            print(f"  no building near node {el['id']} ({el['tags'].get('name', '')})", file=sys.stderr)
            continue
        pt = Point(lng, lat)
        inside = [c for c in cands if c["_poly"].contains(pt)]
        pick = inside[0] if inside else min(cands, key=lambda c: c["_poly"].distance(pt))
        claim(key, pick)
    area_cache: dict[int, list[dict]] = {}
    for key, el in relations:
        if el["id"] not in area_cache:
            area_cache[el["id"]] = fetch_buildings_in_relation(el["id"])
        for b in area_cache[el["id"]]:
            rec = building_record(b, lat0)
            if rec and rec["kind"] not in FOREIGN_BUILDING_KINDS:
                claim(key, rec)
        print(f"  relation {el['id']} ({el['tags'].get('name', '')[:20]}): {len(assigned.get(key, []))} buildings")

    # Every claimed footprint is now re-cut against the basemap's own building
    # parts (the tiles the map draws): whatever the basemap renders inside an
    # OSM building outline — podium + tower parts, integer-rounded heights —
    # becomes our geometry, so a coloured block is never shorter or narrower
    # than the grey one under it. OSM only decides which buildings belong to
    # which school. Campuses with no OSM building way at all are cut directly
    # from the tiles by their outline; if the basemap draws nothing there
    # either, a low slab of the outline keeps the school visible. Campuses
    # whose buildings were claimed by a sibling school sharing the site are
    # left alone.
    def has_claimed_building_inside(poly) -> bool:
        return any(rec["_poly"].representative_point().within(poly) for rec in claimed.values())

    fallback = [(key, poly) for key, poly in polygons if not assigned.get(key) and not has_claimed_building_inside(poly)]
    tiles: set[tuple[int, int]] = set()
    for geom in [rec["_poly"] for recs in assigned.values() for rec in recs] + [poly for _, poly in fallback]:
        minx, miny, maxx, maxy = geom.bounds
        x0, y0 = tile_xy(minx, maxy, TILE_ZOOM)
        x1, y1 = tile_xy(maxx, miny, TILE_ZOOM)
        for tx in range(x0, x1 + 1):
            for ty in range(y0, y1 + 1):
                tiles.add((tx, ty))
    print(f"Re-cutting footprints against basemap tiles ({len(tiles)} tiles; {len(fallback)} campuses need fallback footprints)")
    parts = fetch_tile_building_parts(tiles)

    # Coarse spatial index (~50 m cells) on part representative points.
    cell = 0.0005
    grid: dict[tuple[int, int], list[dict]] = {}
    for part in parts:
        c = part["_poly"].representative_point()
        grid.setdefault((int(c.x / cell), int(c.y / cell)), []).append(part)

    used_parts: set[str] = set()

    def parts_within(poly) -> list[dict]:
        """Basemap parts that mostly (≥ 50 % of their area) lie inside `poly`.
        Area overlap, not point-in-polygon: the z14 tile quantises footprints
        by ~0.6 m and tower parts can spill past the OSM outline."""
        minx, miny, maxx, maxy = poly.bounds
        found = []
        for gx in range(int(minx / cell) - 2, int(maxx / cell) + 3):
            for gy in range(int(miny / cell) - 2, int(maxy / cell) + 3):
                for part in grid.get((gx, gy), []):
                    if part["id"] in used_parts:
                        continue
                    ppoly = part["_poly"]
                    if not ppoly.intersects(poly):
                        continue
                    if ppoly.intersection(poly).area >= 0.5 * ppoly.area:
                        found.append(part)
        for part in found:
            used_parts.add(part["id"])
        return found

    def part_record(part: dict, osm_id: str, name: str | None, kind: str) -> dict:
        return {
            "osmId": osm_id,
            "name": name,
            "kind": kind,
            "height": round(part["height"] + EXTRA_HEIGHT_M, 1),
            "minHeight": round(part["minHeight"], 1),
            "coordinates": buffered_footprint(part["_poly"], lat0),
            "_poly": part["_poly"],
        }

    recut = 0
    for key, recs in assigned.items():
        new_recs = []
        for rec in recs:
            inside = parts_within(rec["_poly"])
            if not inside:
                new_recs.append(rec)  # nothing quantised into this outline: keep the OSM shape
                continue
            recut += 1
            for i, part in enumerate(inside):
                new_recs.append(part_record(part, f"{rec['osmId']}#p{i}" if i else rec["osmId"], rec.get("name"), rec.get("kind") or "yes"))
        assigned[key] = new_recs
    print(f"  {recut} OSM footprints replaced by the basemap's parts")

    for key, poly in fallback:
        inside = parts_within(poly)
        if inside:
            for part in inside:
                claim(key, part_record(part, part["id"], None, "tile"))
            print(f"  {key}: {len(inside)} basemap building parts")
            continue
        el = next(e for k, e in owners if k == key and e["type"] != "node")
        h, mh = parse_height(el["tags"])
        geoms = list(poly.geoms) if poly.geom_type == "MultiPolygon" else [poly]
        for i, g in enumerate(geoms):
            claim(key, {
                "osmId": f"{el['type'][0]}{el['id']}" + (f"#{i}" if i else ""),
                "name": None,
                "kind": "outline",
                "height": round(h + EXTRA_HEIGHT_M, 1),
                "minHeight": mh,
                "coordinates": buffered_footprint(g, lat0),
                "_poly": g,
            })

    # --- assemble output -------------------------------------------------------
    schools = []
    for s in dsedj:
        key = f"dsedj:{s['code']}"
        els = by_code.get(s["code"], [])
        blds = assigned.get(key, [])
        if not els:
            continue
        # Representative point = the best-matched campus (exact/alias before
        # substring), not the mean of all campuses, which for a school with a
        # Coloane annex would land in the sea.
        def match_score(el: dict) -> float:
            return max((m["score"] for m in el.get("_match", []) if m["code"] == s["code"]), default=0.0)

        main = max(els, key=match_score)
        schools.append(
            {
                "id": key,
                "name": {"zh": s["nameZh"], "pt": s["namePt"]},
                "level": s["level"],
                "levels": s["levels"],
                "system": s["system"],
                "coordinates": [round(main["_point"][0], 6), round(main["_point"][1], 6)],
                "osm": [f"{el['type'][0]}{el['id']}" for el in els],
                "buildings": [{k: v for k, v in b.items() if not k.startswith("_")} for b in blds],
            }
        )
    for el in tertiary:
        key = f"osm:{el['type'][0]}{el['id']}"
        blds = assigned.get(key, [])
        tags = el["tags"]
        name = tags.get("name", "")
        schools.append(
            {
                "id": key,
                "name": {"zh": tags.get("name:zh") or zh_part(name) or name, "pt": tags.get("name:pt") or tags.get("name:en") or latin_part(name)},
                "level": "university",
                "levels": {"kindergarten": False, "primary": False, "secondary": False},
                "system": "tertiary",
                "coordinates": [round(el["_point"][0], 6), round(el["_point"][1], 6)],
                "osm": [key.split(":")[1]],
                "buildings": [{k: v for k, v in b.items() if not k.startswith("_")} for b in blds],
            }
        )

    matched_codes = {s["code"] for s in dsedj if by_code.get(s["code"])}
    unmatched_dsedj = [{"code": s["code"], "name": s["nameZh"], "level": s["level"]} for s in dsedj if s["code"] not in matched_codes]
    no_buildings = [s["name"]["zh"] for s in schools if not s["buildings"]]

    if len(schools) < MIN_SCHOOLS:
        print(f"ERROR: only {len(schools)} schools resolved (< {MIN_SCHOOLS}); refusing to write", file=sys.stderr)
        return 1

    output = {
        "fetchedAtUtc": datetime.now(tz=timezone.utc).isoformat(),
        "sources": {
            "dsedj": DSEDJ_PAGE,
            "osm": "OpenStreetMap via Overpass (amenity=school/kindergarten/college/university + building ways)",
        },
        "levels": list(LEVELS),
        "schools": schools,
        "unmatchedDsedj": unmatched_dsedj,
        # Named drops only: unnamed building=school outlines that matched no
        # school are usually annexes already covered through their campus.
        "droppedOsm": sorted({(el["tags"].get("amenity") or el["tags"].get("building") or "") + ":" + el["tags"]["name"] for el in dropped if el["tags"].get("name")}),
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    by_level = {}
    for s in schools:
        by_level[s["level"]] = by_level.get(s["level"], 0) + 1
    total_b = sum(len(s["buildings"]) for s in schools)
    print(f"\nDone. {len(schools)} schools/campuses, {total_b} buildings; by level {by_level}")
    print(f"DSEDJ schools without an OSM match: {len(unmatched_dsedj)} -> {[u['name'] for u in unmatched_dsedj]}")
    print(f"Schools matched but with no building footprint: {len(no_buildings)} -> {no_buildings}")
    print(f"Wrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
